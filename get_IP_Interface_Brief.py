import sys, time, paramiko, json, re, os, multiprocessing, concurrent.futures
from netmiko import Netmiko
from netmiko.exceptions import NetmikoAuthenticationException
from openpyxl import Workbook
from openpyxl.styles import Alignment
from getInputs import getInputs

try: os.mkdir("scripts_Outputs")
except FileExistsError: pass

credentials = getInputs(f'{os.getcwd()}\\scripts_Outputs\\', os.path.basename(__file__)[:-3])


def autoFit_sheet(sheet):
	for column in sheet.columns:
		max_length = 0
		column_letter = column[0].column_letter
		for cell in column:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except:
				pass
		adjusted_width = (max_length + 4)
		sheet.column_dimensions[column_letter].width = adjusted_width

def CenterCells(sheet, num_of_columns):
	for i in range(1, num_of_columns+1):
		sheet.cell(row=sheet.max_row, column=i).alignment = Alignment(horizontal='center', vertical='center')


def generate_excel():
	folders = [folder for folder in os.listdir('.') if os.path.isdir(folder)]

	wb = Workbook()
	files_list = []
	for folder in folders:
		files_list += [f"{folder}/{i}" for i in os.listdir(folder) if i.endswith(".txt")]

	interfacesSheet = wb.create_sheet("disp ip interface brief", 0)
	x = wb.remove(wb["Sheet"])
	Columns_headers = ['NE Name', 'Interface', 'IP Address', 'Mask', 'Physical', 'Protocol', 'VPN'] #########
	interfacesSheet.append(Columns_headers)
	CenterCells(interfacesSheet, len(Columns_headers))
	for site_file in files_list:
		node_name = site_file.split('/')[-1][:-4]
		with open(site_file, 'r') as f:
			txt = f.read()
		strict_info_regex = r'(Interface +IP Address.+?)\S*<{}>'.format(node_name) #########
		site_info = re.split(r'[\r\n]+', re.findall(strict_info_regex, txt, re.DOTALL)[0])
		interfaces = [i.strip().split() for i in site_info[1:] if i != '']
		for i in interfaces:
			if i[1] == 'unassigned':
				interfacesSheet.append([node_name, i[0], 'unassigned', 'unassigned'] + i[2:])
			else:
				interfacesSheet.append([node_name, i[0]] + i[1].split('/') + i[2:])
	autoFit_sheet(interfacesSheet)

	excel_file_name = f'disp_ip_interface_brief_{time.strftime("%H;%M;%S")}'
	wb.save(f'{excel_file_name}.xlsx')
	print(f"Excel file {excel_file_name}.xlsx created successfully!")


start_time = time.time()
###### Using Paramiko


counter = 0
def mainCode(remote, counter):
	if remote["host"] == "" or remote["username"] == "" or remote["password"] == "": return
	try: 
		host = {
		"host": remote["host"],
		"username": remote["username"],
		"password": remote["password"],
		"device_type": "huawei_vrp",
		"session_log": f'Logs/{remote["NE_Name"]}_session.log',
		}#, "global_delay_factor": 2}
		shell = Netmiko(**host)
		time.sleep(1)
	# except NetmikoAuthenticationException as e:
	# # Handle authentication errors
		# print(f'Authentication error for: {remote["host"]}, Pass is not {remote["password"]}')
		# return
	except Exception as e:
		# Handle unexpected exceptions
		print(f"An unexpected error occurred while trying to connect to {remote["NE_Name"]}: {str(e)}")
		return 0
	try:
		output = ''
		device_Sysname = shell.find_prompt().split('<')[-1].strip('>')
		command = "display ip interface brief"
		output += shell.send_command_timing(command, strip_prompt=False, strip_command=False, read_timeout=0)
		print(f"Device name: {device_Sysname}, Success")
		counter += 1
	except Exception as e:
		# Handle unexpected exceptions
		print(f"An unexpected error occurred while executing commands on {remote["NE_Name"]}: {str(e)}")
		return 0
	# shell.disconnect()
	# current_time = time.strftime("%H;%M;%S")
	# current_date = time.strftime("%Y-%m-%d")
	
	


	filename = device_Sysname+'.txt' #+ "\\" + "Commit_changes_with_username_" + current_date + "_" + current_time + ".txt"
	with open(remote['Foldername'] + '/' + filename, 'w') as f:
		# f.write(re.sub(r'  ---- More ----\S+\s+\S+16D', '', output[output.index(command):]))
		f.write(output)
	return counter


num_threads = 100#min(int(multiprocessing.cpu_count() * (3/4)), len(credentials))

# Create a thread pool executor with the desired number of threads
with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
	# Submit the slow function to the thread pool executor for each piece of data
	futures = [executor.submit(mainCode, remote, counter) for remote in credentials]
	# Wait for all the threads to complete
	results = [f.result() for f in futures]
	sum_counter = sum(results)


end_time = time.time()
print(f"success rate = {sum_counter}/{len(credentials)}")
# Calculate and print the elapsed time
execution_time = end_time - start_time
print(f"Execution time: {execution_time:.6f} seconds")

print("Generating Excel File...")
generate_excel()