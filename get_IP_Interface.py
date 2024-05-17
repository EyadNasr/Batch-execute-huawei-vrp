import sys, time, paramiko, json, re, os, multiprocessing, concurrent.futures, textfsm
from netmiko import Netmiko
from netmiko.exceptions import NetmikoAuthenticationException
from openpyxl import Workbook
from openpyxl.styles import Alignment
from getInputs import getInputs

try: os.mkdir("scripts_Outputs")
except FileExistsError: pass

credentials = getInputs(f'{os.getcwd()}\\scripts_Outputs\\', os.path.basename(__file__)[:-3])


def autoFit_sheet(sheet):
	for column in sheet.columns:
		max_length = 0
		column_letter = column[0].column_letter
		for cell in column:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except:
				pass
		adjusted_width = (max_length + 4)
		sheet.column_dimensions[column_letter].width = adjusted_width

def CenterCells(sheet, num_of_columns):
	for i in range(1, num_of_columns+1):
		sheet.cell(row=sheet.max_row, column=i).alignment = Alignment(horizontal='center', vertical='center')

def parse_router_output(template_file, router_output):
    with open(template_file) as template:
        fsm_table = textfsm.TextFSM(template)
        result = fsm_table.ParseText(router_output)
        headers = fsm_table.header

        # Combine headers and parsed data into a dictionary
        parsed_data = [dict(zip(headers, entry)) for entry in result]

        return parsed_data

def generate_excel():
	template_file = '..\\..\\textFSM_templates\\display_ip_interface.template'

	wb = Workbook()
	files_list = []
	folders = [folder for folder in os.listdir('.') if os.path.isdir(folder)]
	for folder in folders:
		files_list += [f"{folder}/{i}" for i in os.listdir(folder) if i.endswith(".txt")]

	interfacesSheet = wb.create_sheet("disp ip interface", 0)
	x = wb.remove(wb["Sheet"])
	with open(template_file) as template:
		fsm_table = textfsm.TextFSM(template)
		Columns_headers = fsm_table.header
	interfacesSheet.append(['NE Name'] + Columns_headers)
	CenterCells(interfacesSheet, len(Columns_headers))

	for site_file in files_list:
		node_name = site_file.split('/')[-1][:-4]
		with open(site_file, 'r') as f:
			router_output = '\n'.join(re.split('[\r\n]+', f.read()))
		parsed_data = parse_router_output(template_file, router_output)
		for entry in parsed_data:
			if not all(value == '' for value in list(entry.values())[:-1]):
				if entry['IP_PROCESSING_STATE'] == '': entry['IP_PROCESSING_STATE'] = 'Enabled'
				interfacesSheet.append([node_name] + list(entry.values()))
	autoFit_sheet(interfacesSheet)

	excel_file_name = 'disp_ip_interface'
	wb.save(f'{excel_file_name}.xlsx')
	print(f"Excel file {excel_file_name}.xlsx created successfully!")



start_time = time.time()
###### Using Netmiko


counter = 0
def mainCode(remote, counter):
	if remote["host"] == "" or remote["username"] == "" or remote["password"] == "": return
	try: 
		host = {
		"host": remote["host"],
		"username": remote["username"],
		"password": remote["password"],
		"device_type": "huawei_vrp",
		"session_log": f'Logs/{remote["NE_Name"]}_session.log',
		}#, "global_delay_factor": 2}
		shell = Netmiko(**host)
		time.sleep(1)
	# except NetmikoAuthenticationException as e:
	# # Handle authentication errors
		# print(f'Authentication error for: {remote["host"]}, Pass is not {remote["password"]}')
		# return
	except Exception as e:
		# Handle unexpected exceptions
		print(f"An unexpected error occurred while trying to connect: {str(e)}")
		return 0
	try:
		output = ''
		device_Sysname = shell.find_prompt().split('<')[-1].strip('>')
		command = 'display ip interface'
		output += shell.send_command_timing(command, strip_prompt=False, strip_command=False, read_timeout=0)
		print(f"Device name: {device_Sysname}, Success")
		counter += 1
	except Exception as e:
		# Handle unexpected exceptions
		print(f"An unexpected error occurred: {str(e)}", remote)
		return 0
	# shell.disconnect()
	# current_time = time.strftime("%H;%M;%S")
	# current_date = time.strftime("%Y-%m-%d")
	
	


	filename = device_Sysname+'.txt' #+ "\\" + "Commit_changes_with_username_" + current_date + "_" + current_time + ".txt"
	with open(remote['Foldername'] + '/' + filename, 'w') as f:
		# f.write(re.sub(r'  ---- More ----\S+\s+\S+16D', '', output[output.index(command):]))
		f.write(output)#[labels.index(command):])
	return counter


num_threads = 100#min(int(multiprocessing.cpu_count() * (3/4)), len(credentials))

# Create a thread pool executor with the desired number of threads
with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
	# Submit the slow function to the thread pool executor for each piece of data
	futures = [executor.submit(mainCode, remote, counter) for remote in credentials]
	# Wait for all the threads to complete
	results = [f.result() for f in futures]
	sum_counter = sum(results)


end_time = time.time()
print(f"success rate = {sum_counter}/{len(credentials)}")
# Calculate and print the elapsed time
execution_time = end_time - start_time
print(f"Execution time: {execution_time:.6f} seconds")

print("Generating Excel File...")
generate_excel()