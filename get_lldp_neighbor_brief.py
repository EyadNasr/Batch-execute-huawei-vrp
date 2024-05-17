import sys, time, paramiko, json, re, os, multiprocessing, concurrent.futures, textfsm
from netmiko import Netmiko
from netmiko.exceptions import NetmikoAuthenticationException
from openpyxl import Workbook
from openpyxl.styles import Alignment
from getInputs import getInputs

try: os.mkdir("scripts_Outputs")
except FileExistsError: pass

credentials = getInputs(f'{os.getcwd()}\\scripts_Outputs\\', os.path.basename(__file__)[:-3])

def autoFit_sheet(sheet):
	for column in sheet.columns:
		max_length = 0
		column_letter = column[0].column_letter
		for cell in column:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except:
				pass
		adjusted_width = (max_length + 4)
		sheet.column_dimensions[column_letter].width = adjusted_width

def CenterCells(sheet, num_of_columns):
	for i in range(1, num_of_columns+1):
		sheet.cell(row=sheet.max_row, column=i).alignment = Alignment(horizontal='center', vertical='center')

def parse_router_output(template_file, router_output):
    with open(template_file) as template:
        fsm_table = textfsm.TextFSM(template)
        result = fsm_table.ParseText(router_output)
        headers = fsm_table.header

        # Combine headers and parsed data into a dictionary
        parsed_data = [dict(zip(headers, entry)) for entry in result]

        return parsed_data

def get_rows(node_name, regex, txt, template_file):
	strict_info_regex = regex.format(node_name) #########
	router_output = re.findall(strict_info_regex, txt, re.DOTALL)[0]
	parsed_data = parse_router_output(template_file, router_output)
	return parsed_data

def generate_excel():
	folders = [folder for folder in os.listdir('.') if os.path.isdir(folder) and not folder.startswith('Logs')]
	files_list = []
	for folder in folders:
		files_list += [f"{folder}/{i}" for i in os.listdir(folder) if i.endswith(".txt")]

	wb = Workbook()
	interfacesSheet = wb.create_sheet("disp lldp neighbor brief", 0)
	x = wb.remove(wb["Sheet"])
	Columns_headers = ['NE Name', 'Local Interface', 'Neighbor Device', 'Neighbor Interface', 'Exptime (sec)'] #########
	interfacesSheet.append(Columns_headers)
	CenterCells(interfacesSheet, len(Columns_headers))

	LocalIntf_regex = re.compile(r'Local Intf.+?<\S+>', re.DOTALL)
	LocalInterface_regex = re.compile(r'Local Interface.+?<\S+>', re.DOTALL)
	parent = '..\\..\\textFSM_templates\\'
	template_file1 = f'{parent}display_lldp_neighbor_brief1.template'
	template_file2 = f'{parent}display_lldp_neighbor_brief2.template'

	for site_file in files_list:
		node_name = site_file.split('/')[-1][:-4]
		with open(site_file, 'r') as f:
			txt = f.read()
		if re.search(LocalIntf_regex, txt):
			regex = r'Local Intf +Neighbor Dev[^\r\n]+[\r\n]+-*[\r\n]*?(.*?)[\r\n]+\S*<{}>'
			parsed_data = get_rows(node_name, regex, txt, template_file1)
		elif re.search(LocalInterface_regex, txt):
			regex = r'Local Interface +Exptime[^\r\n]+[\r\n]+-*[\r\n]*?(.*?)[\r\n]+\S*<{}>'
			parsed_data = get_rows(node_name, regex, txt, template_file2)
		else: 
			print(node_name, 'Not matched')
			continue
		
		for entry in parsed_data:
			interfacesSheet.append([node_name] + list(entry.values()))

	autoFit_sheet(interfacesSheet)

	excel_file_name = 'disp_lldp_neighbor_brief'
	wb.save(f'{excel_file_name}.xlsx')
	print(f"Excel file {excel_file_name}.xlsx created successfully!")

start_time = time.time()
###### Using Paramiko


counter = 0
def mainCode(remote, counter):
	if remote["host"] == "" or remote["username"] == "" or remote["password"] == "": return
	try: 
		host = {
		"host": remote["host"],
		"username": remote["username"],
		"password": remote["password"],
		"device_type": "huawei_vrp",
		"session_log": f'Logs/{remote["NE_Name"]}_session.log',
		}#, "global_delay_factor": 2}
		shell = Netmiko(**host)
		time.sleep(1)
	# except NetmikoAuthenticationException as e:
	# # Handle authentication errors
		# print(f'Authentication error for: {remote["host"]}, Pass is not {remote["password"]}')
		# return
	except Exception as e:
		# Handle unexpected exceptions
		print(f"An unexpected error occurred while trying to connect to {remote["NE_Name"]}: {str(e)}")
		return 0
	try:
		output = ''
		device_Sysname = shell.find_prompt().split('<')[-1].strip('>')
		command = "display lldp neighbor brief"
		output += shell.send_command_timing(command, strip_prompt=False, strip_command=False, read_timeout=0)
		print(f"Device name: {device_Sysname}, Success")
		counter += 1
	except Exception as e:
		# Handle unexpected exceptions
		print(f"An unexpected error occurred while executing commands on {remote["NE_Name"]}: {str(e)}")
		return 0
	# shell.disconnect()
	# current_time = time.strftime("%H;%M;%S")
	# current_date = time.strftime("%Y-%m-%d")
	
	


	filename = device_Sysname+'.txt' #+ "\\" + "Commit_changes_with_username_" + current_date + "_" + current_time + ".txt"
	with open(remote['Foldername'] + '/' + filename, 'w') as f:
		# f.write(re.sub(r'  ---- More ----\S+\s+\S+16D', '', output[output.index(command):]))
		f.write(output)
	return counter


num_threads = 100#min(int(multiprocessing.cpu_count() * (3/4)), len(credentials))

# Create a thread pool executor with the desired number of threads
with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
	# Submit the slow function to the thread pool executor for each piece of data
	futures = [executor.submit(mainCode, remote, counter) for remote in credentials]
	# Wait for all the threads to complete
	results = [f.result() for f in futures]
	sum_counter = sum(results)


end_time = time.time()
print(f"success rate = {sum_counter}/{len(credentials)}")
# Calculate and print the elapsed time
execution_time = end_time - start_time
print(f"Execution time: {execution_time:.6f} seconds")

print("Generating Excel File...")
generate_excel()